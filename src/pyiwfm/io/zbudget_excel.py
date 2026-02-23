"""
Zone budget Excel export for IWFM post-processing.

Generates Excel workbooks from IWFM zone-budget HDF5 data with one sheet
per zone, title lines, bold headers, and auto-fitted column widths.
"""

from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from pyiwfm.io.budget_utils import filter_time_range
from pyiwfm.io.zbudget import ZBudgetReader
from pyiwfm.io.zbudget_control import ZBudgetControlConfig

logger = logging.getLogger(__name__)


def _make_sheet_name(name: str) -> str:
    """Create a valid Excel sheet name (max 31 chars, no invalid chars)."""
    invalid = r"[]:*?/\\"
    clean = name
    for ch in invalid:
        clean = clean.replace(ch, "_")
    return clean[:31]


def zbudget_to_excel(
    reader: ZBudgetReader,
    output_path: Path | str,
    area_factor: float = 1.0,
    volume_factor: float = 1.0,
    area_unit: str = "AC",
    volume_unit: str = "AC.FT.",
    zone_ids: list[int] | None = None,
    zone_def_file: Path | str | None = None,
    begin_date: str | None = None,
    end_date: str | None = None,
    output_interval: str | None = None,
    layer: int = 1,
) -> Path:
    """Export zone budget data to an Excel workbook.

    One sheet per zone.  Same formatting as budget export:
    title area (bold), column headers (bold), data rows, auto-fit widths.

    Parameters
    ----------
    reader : ZBudgetReader
        Open zone-budget reader.
    output_path : Path or str
        Destination ``.xlsx`` file.
    area_factor, volume_factor : float
        FACTAROU / FACTVLOU from the control file.
    area_unit, volume_unit : str
        UNITAROU / UNITVLOU for title substitution.
    zone_ids : list[int] or None
        1-based zone IDs to include.  ``None`` or ``[-1]`` → all.
    zone_def_file : Path or str or None
        External zone definition file (not yet used — reserved).
    begin_date, end_date : str or None
        IWFM datetime strings for time filtering.
    output_interval : str or None
        Reserved for future resampling.
    layer : int
        Model layer (1-based).

    Returns
    -------
    Path
        The written workbook path.
    """
    output_path = Path(output_path)

    # Determine which zones to export
    all_zones = reader.zones
    if zone_ids is None or zone_ids == [-1]:
        zone_names = all_zones
    elif not zone_ids:
        zone_names = []
    else:
        zone_names = []
        for zid in zone_ids:
            if 1 <= zid <= len(all_zones):
                zone_names.append(all_zones[zid - 1])

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    bold_font = Font(bold=True)

    for zone_name in zone_names:
        sheet_name = _make_sheet_name(zone_name)
        ws = wb.create_sheet(title=sheet_name)

        # --- Title area ---
        row_num = 1
        title = f"{reader.descriptor} — {zone_name}"
        ws.cell(row=row_num, column=1, value=title).font = bold_font
        row_num += 1

        zone_info = reader.get_zone_info(zone_name)
        if zone_info.area > 0:
            area_line = f"Area: {zone_info.area * area_factor:.2f} {area_unit}"
            ws.cell(row=row_num, column=1, value=area_line).font = bold_font
            row_num += 1

        unit_line = f"Volume Unit: {volume_unit}"
        ws.cell(row=row_num, column=1, value=unit_line).font = bold_font
        row_num += 1

        # Blank row
        row_num += 1

        # --- Column headers ---
        col_names = reader.data_names
        all_headers = ["Time"] + col_names
        for col_idx_h, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx_h, value=header)
            cell.font = bold_font
        row_num += 1

        # --- Data rows (with unit conversion) ---
        df = reader.get_dataframe(zone_name, layer=layer, volume_factor=volume_factor)
        df = filter_time_range(df, begin_date, end_date)

        if isinstance(df.index, pd.DatetimeIndex):
            time_strings = [dt.strftime("%m/%d/%Y_%H:%M") for dt in df.index]
        else:
            time_strings = [str(t) for t in df.index]

        values = df.to_numpy(dtype=np.float64, na_value=np.nan)
        for i in range(len(df)):
            ws.cell(row=row_num, column=1, value=time_strings[i])
            for j in range(df.shape[1]):
                num = values[i, j]
                ws.cell(
                    row=row_num,
                    column=j + 2,
                    value=round(float(num), 4) if not np.isnan(num) else None,
                )
            row_num += 1

        # --- Auto-fit column widths ---
        for col_idx_w in range(1, len(all_headers) + 1):
            col_letter = get_column_letter(col_idx_w)
            max_width = 0
            for row_cells in ws.iter_rows(
                min_col=col_idx_w, max_col=col_idx_w, min_row=1, max_row=row_num - 1
            ):
                for cell in row_cells:
                    if cell.value is not None:
                        max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    if not wb.sheetnames:
        wb.create_sheet(title="(no data)")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def zbudget_control_to_excel(config: ZBudgetControlConfig) -> list[Path]:
    """Process a full zbudget control file, generating one ``.xlsx`` per spec.

    Parameters
    ----------
    config : ZBudgetControlConfig
        Parsed zone-budget control configuration.

    Returns
    -------
    list[Path]
        Paths of created workbook files.
    """
    created: list[Path] = []

    for spec in config.zbudgets:
        if not spec.hdf_file.exists():
            logger.warning("ZBudget HDF5 file not found: %s", spec.hdf_file)
            continue

        reader = ZBudgetReader(spec.hdf_file)

        out = spec.output_file
        if out.suffix.lower() not in (".xlsx", ".xls"):
            out = out.with_suffix(".xlsx")

        result = zbudget_to_excel(
            reader=reader,
            output_path=out,
            area_factor=config.area_factor,
            volume_factor=config.volume_factor,
            area_unit=config.area_unit,
            volume_unit=config.volume_unit,
            zone_ids=spec.zone_ids if spec.zone_ids else None,
            zone_def_file=spec.zone_def_file,
            begin_date=config.begin_date,
            end_date=config.end_date,
            output_interval=spec.output_interval,
        )
        created.append(result)
        logger.info("Wrote %s (%d zones)", result, reader.n_zones)

    return created
