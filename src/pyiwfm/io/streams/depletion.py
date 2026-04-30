"""Stream depletion analysis — compare streamflow between baseline and pumping scenarios.

**Read-only analysis.** Computes depletion by comparing two model runs;
results are a *report*, not a model input. Tabular, plot, and map
exports are added in Phase 2.2.a. See
``docs/user_guide/inputs_vs_outputs.rst``.

Two granularities are supported:

- :func:`compute_stream_depletion` — per-reach depletion read from the
  stream **reach** budget HDF declared in each model's stream main file
  (key ``stream_budget_file`` in :attr:`IWFMModel.metadata`).
- :func:`compute_stream_node_depletion` — per-stream-node depletion read
  from the stream **node** budget HDF (key ``stream_node_budget_file``).

Both look up the budget filename declared in each model's input rather
than globbing, and both raise a clear :class:`BudgetOutputMissingError`
if either model didn't declare the required output or if the file
doesn't exist on disk. Column matching is **exact** against the
``sa_column`` argument (default ``"Stream-Aquifer Interaction Within Model"``).
"""

from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import TYPE_CHECKING, Literal

import numpy as np

if TYPE_CHECKING:
    from numpy.typing import NDArray

    from pyiwfm.core.model import IWFMModel


# Canonical IWFM column header for stream-aquifer interaction. Both the
# stream reach budget and stream node budget use this exact name in
# IWFM v5+. Older models may use variants like ``"Gain from GW (+)"``;
# pass that explicitly via the ``sa_column`` parameter.
DEFAULT_SA_COLUMN: str = "Stream-Aquifer Interaction Within Model"


class BudgetOutputMissingError(ValueError):
    """Raised when a model required for depletion analysis didn't declare
    or didn't produce a needed budget output file.

    Carries the model label (``"baseline"`` / ``"scenario"``), the kind of
    budget (``"reach"`` / ``"node"``), the looked-up metadata key, and
    the reason the lookup failed. Use the message in user-facing CLI
    output; it tells the operator how to fix their input deck.
    """

    def __init__(
        self,
        model_label: str,
        kind: Literal["reach", "node"],
        metadata_key: str,
        reason: str,
    ) -> None:
        self.model_label = model_label
        self.kind = kind
        self.metadata_key = metadata_key
        self.reason = reason
        kind_pretty = "stream reach budget" if kind == "reach" else "stream node budget"
        super().__init__(
            f"{model_label} model has no usable {kind_pretty}: {reason}. "
            f"Enable this output in the model's stream main file (the "
            f"{metadata_key!r} entry in IWFMModel.metadata is populated "
            f"during loading from STRMRCHBUDFL / STNDBUDFL respectively)."
        )


@dataclasses.dataclass
class StreamDepletionResult:
    """Result of stream depletion analysis at a single reach/node."""

    reach_id: int
    reach_name: str
    times: list[str]
    baseline_flow: NDArray[np.float64]
    scenario_flow: NDArray[np.float64]
    depletion: NDArray[np.float64]  # baseline - scenario (positive = depletion)
    cumulative_depletion: NDArray[np.float64]
    max_depletion: float
    max_depletion_timestep: int
    total_depletion: float  # cumulative at last timestep

    def to_dict(self) -> dict[str, object]:
        """Serialize for API response."""
        return {
            "reach_id": self.reach_id,
            "reach_name": self.reach_name,
            "times": self.times,
            "depletion": [round(float(v), 4) for v in self.depletion],
            "cumulative_depletion": [round(float(v), 4) for v in self.cumulative_depletion],
            "max_depletion": round(float(self.max_depletion), 4),
            "max_depletion_timestep": int(self.max_depletion_timestep),
            "total_depletion": round(float(self.total_depletion), 4),
        }


@dataclasses.dataclass
class StreamDepletionReport:
    """Aggregate depletion results across multiple reaches."""

    results: list[StreamDepletionResult]
    n_reaches: int
    n_timesteps: int
    total_max_depletion: float  # max depletion across all reaches and timesteps
    total_cumulative_depletion: float  # sum of cumulative depletion at last timestep

    def to_dict(self) -> dict[str, object]:
        """Serialize for API response."""
        return {
            "n_reaches": self.n_reaches,
            "n_timesteps": self.n_timesteps,
            "total_max_depletion": round(float(self.total_max_depletion), 4),
            "total_cumulative_depletion": round(float(self.total_cumulative_depletion), 4),
            "reaches": [r.to_dict() for r in self.results],
        }


@dataclasses.dataclass
class StreamNodeDepletionResult:
    """Per-stream-node depletion result.

    Attributes
    ----------
    stream_node_id
        1-based stream node ID (matches ``model.streams.nodes`` keys).
    times
        Timestamp strings, one per element of the flux arrays.
    baseline_sa_flux, scenario_sa_flux
        Stream-aquifer interaction flux at this node, baseline vs.
        scenario. Sign convention follows the IWFM column verbatim
        (typically positive = stream gain from aquifer, negative =
        stream loss to aquifer).
    depletion
        ``baseline_sa_flux - scenario_sa_flux``. Positive means the
        scenario reduced gain (or increased loss) — i.e. depletion.
    cumulative_depletion
        Running sum of ``depletion``.
    max_depletion
        Largest absolute depletion across all timesteps.
    max_depletion_timestep
        0-based timestep index where ``max_depletion`` occurs.
    total_depletion
        ``cumulative_depletion[-1]``.
    """

    stream_node_id: int
    times: list[str]
    baseline_sa_flux: NDArray[np.float64]
    scenario_sa_flux: NDArray[np.float64]
    depletion: NDArray[np.float64]
    cumulative_depletion: NDArray[np.float64]
    max_depletion: float
    max_depletion_timestep: int
    total_depletion: float

    def to_dict(self) -> dict[str, object]:
        """Serialize for API responses / JSON export."""
        return {
            "stream_node_id": self.stream_node_id,
            "times": self.times,
            "depletion": [round(float(v), 4) for v in self.depletion],
            "cumulative_depletion": [round(float(v), 4) for v in self.cumulative_depletion],
            "max_depletion": round(float(self.max_depletion), 4),
            "max_depletion_timestep": int(self.max_depletion_timestep),
            "total_depletion": round(float(self.total_depletion), 4),
        }


@dataclasses.dataclass
class StreamNodeDepletionReport:
    """Aggregate per-stream-node depletion results.

    Companion to :class:`StreamDepletionReport` but at finer granularity:
    one entry per stream node instead of one per reach. Use this when
    you need spatial maps that resolve depletion within a reach (e.g.
    a longitudinal profile, or a point-scatter map of which stream
    nodes are most affected by a pumping scenario).
    """

    results: list[StreamNodeDepletionResult]
    n_stream_nodes: int
    n_timesteps: int
    total_max_depletion: float
    total_cumulative_depletion: float

    def to_dict(self) -> dict[str, object]:
        """Serialize for API responses / JSON export."""
        return {
            "n_stream_nodes": self.n_stream_nodes,
            "n_timesteps": self.n_timesteps,
            "total_max_depletion": round(float(self.total_max_depletion), 4),
            "total_cumulative_depletion": round(float(self.total_cumulative_depletion), 4),
            "stream_nodes": [r.to_dict() for r in self.results],
        }


def compute_stream_depletion(
    baseline_dir: str | Path,
    scenario_dir: str | Path,
    reach_ids: list[int] | None = None,
    budget_type: str = "Stream Reach Budgets",
    *,
    sa_column: str = DEFAULT_SA_COLUMN,
) -> StreamDepletionReport:
    """Compute per-**reach** stream depletion by comparing two model runs.

    Searches each results directory for a stream-reach budget HDF
    (``*Stream*Budget*.hdf``) and extracts the named stream-aquifer
    interaction column. Prefer
    :func:`compute_stream_depletion_from_models` when you have loaded
    :class:`IWFMModel` instances — that variant uses the model-declared
    budget filename and gives clearer errors when an output is missing.

    Parameters
    ----------
    baseline_dir
        Path to the baseline (no-pumping) model results directory.
    scenario_dir
        Path to the pumping-scenario model results directory.
    reach_ids
        Specific 1-based reach IDs to analyze. ``None`` (default) analyzes
        all reaches.
    budget_type
        Reserved for future multi-budget disambiguation.
    sa_column
        Exact column header for the stream-aquifer interaction flux.
        Defaults to :data:`DEFAULT_SA_COLUMN` (the IWFM v5+ canonical
        name). For older models that emit ``"Gain from GW (+)"``, pass
        that name explicitly.

    Returns
    -------
    StreamDepletionReport
        Per-reach depletion results.

    Raises
    ------
    FileNotFoundError
        If a budget HDF cannot be located in either results directory.
    KeyError
        If ``sa_column`` is not present in the budget headers (the
        message lists the available columns).
    """
    from pyiwfm.io.budget import BudgetReader

    baseline_path = Path(baseline_dir)
    scenario_path = Path(scenario_dir)

    # Find stream budget files
    baseline_budget = _find_stream_budget(baseline_path, budget_type)
    scenario_budget = _find_stream_budget(scenario_path, budget_type)

    baseline_reader = BudgetReader(str(baseline_budget))
    scenario_reader = BudgetReader(str(scenario_budget))

    return _compute_reach_depletion_from_readers(
        baseline_reader,
        scenario_reader,
        reach_ids=reach_ids,
        sa_column=sa_column,
    )


def compute_stream_depletion_from_models(
    baseline_model: IWFMModel,
    scenario_model: IWFMModel,
    *,
    reach_ids: list[int] | None = None,
    sa_column: str = DEFAULT_SA_COLUMN,
) -> StreamDepletionReport:
    """Per-reach stream depletion using model-declared reach budget files.

    Resolves the budget HDF for each model from the
    ``stream_budget_file`` entry in :attr:`IWFMModel.metadata`
    (populated during load from the ``STRMRCHBUDFL`` line of the stream
    main file). Raises :class:`BudgetOutputMissingError` with an
    actionable message if either model didn't declare the output or the
    declared file doesn't exist.

    Parameters
    ----------
    baseline_model
        Loaded baseline (no-pumping) model.
    scenario_model
        Loaded scenario (pumping) model.
    reach_ids
        Specific 1-based reach IDs to analyze. ``None`` analyzes all.
    sa_column
        Exact column header for stream-aquifer interaction. See
        :data:`DEFAULT_SA_COLUMN`.

    Returns
    -------
    StreamDepletionReport

    Raises
    ------
    BudgetOutputMissingError
        Either model didn't declare a stream reach budget output, or the
        declared file doesn't exist on disk.
    KeyError
        ``sa_column`` not present in the budget.
    """
    from pyiwfm.io.budget import BudgetReader

    baseline_budget = _resolve_budget_file_from_model(
        baseline_model, kind="reach", model_label="baseline"
    )
    scenario_budget = _resolve_budget_file_from_model(
        scenario_model, kind="reach", model_label="scenario"
    )
    return _compute_reach_depletion_from_readers(
        BudgetReader(str(baseline_budget)),
        BudgetReader(str(scenario_budget)),
        reach_ids=reach_ids,
        sa_column=sa_column,
    )


def _compute_reach_depletion_from_readers(
    baseline_reader: object,
    scenario_reader: object,
    *,
    reach_ids: list[int] | None,
    sa_column: str,
) -> StreamDepletionReport:
    """Inner driver shared by both reach-level entry points."""
    locations = baseline_reader.locations  # type: ignore[attr-defined]
    if reach_ids is not None:
        location_indices = [i for i in range(len(locations)) if i + 1 in reach_ids]
    else:
        location_indices = list(range(len(locations)))

    results: list[StreamDepletionResult] = []

    for loc_idx in location_indices:
        reach_name = locations[loc_idx] if loc_idx < len(locations) else f"Reach {loc_idx + 1}"

        base_times, base_values = baseline_reader.get_values(loc_idx)  # type: ignore[attr-defined]
        _scen_times, scen_values = scenario_reader.get_values(loc_idx)  # type: ignore[attr-defined]

        headers = baseline_reader.get_column_headers(loc_idx)  # type: ignore[attr-defined]
        base_flow = _extract_named_column(headers, base_values, sa_column)
        scen_flow = _extract_named_column(headers, scen_values, sa_column)

        n_ts = min(len(base_flow), len(scen_flow))
        base_flow = base_flow[:n_ts]
        scen_flow = scen_flow[:n_ts]

        depletion = base_flow - scen_flow
        cumulative = np.cumsum(depletion)
        times = _format_times(base_times[:n_ts], baseline_reader)
        max_dep_idx = int(np.argmax(np.abs(depletion))) if depletion.size else 0

        results.append(
            StreamDepletionResult(
                reach_id=loc_idx + 1,
                reach_name=reach_name,
                times=times,
                baseline_flow=base_flow,
                scenario_flow=scen_flow,
                depletion=depletion,
                cumulative_depletion=cumulative,
                max_depletion=float(np.max(np.abs(depletion))) if depletion.size else 0.0,
                max_depletion_timestep=max_dep_idx,
                total_depletion=float(cumulative[-1]) if cumulative.size else 0.0,
            )
        )

    total_max = max((r.max_depletion for r in results), default=0.0)
    total_cum = sum(r.total_depletion for r in results)
    n_timesteps = results[0].depletion.shape[0] if results else 0

    return StreamDepletionReport(
        results=results,
        n_reaches=len(results),
        n_timesteps=n_timesteps,
        total_max_depletion=total_max,
        total_cumulative_depletion=total_cum,
    )


def compute_stream_node_depletion(
    baseline_model: IWFMModel,
    scenario_model: IWFMModel,
    *,
    node_ids: list[int] | None = None,
    sa_column: str = DEFAULT_SA_COLUMN,
) -> StreamNodeDepletionReport:
    """Per-**stream-node** depletion using model-declared node budget files.

    Reads the stream-aquifer interaction column from each model's stream
    node budget HDF (declared in the model's stream main file as
    ``STNDBUDFL`` and surfaced in :attr:`IWFMModel.metadata` under
    ``"stream_node_budget_file"``). The returned
    :class:`StreamNodeDepletionReport` lets you build node-resolved maps
    and longitudinal profiles within a reach, where the reach-level
    :func:`compute_stream_depletion_from_models` would only show one
    aggregate value per reach.

    Both models must declare the node budget output AND have produced
    the file on disk. If either one is missing, this raises
    :class:`BudgetOutputMissingError` with a clear remediation message.

    Parameters
    ----------
    baseline_model
        Loaded baseline (no-pumping) model.
    scenario_model
        Loaded scenario (pumping) model.
    node_ids
        Specific 1-based stream node IDs to analyze. ``None`` (default)
        analyzes every stream node in the budget.
    sa_column
        Exact column header for stream-aquifer interaction. Default
        :data:`DEFAULT_SA_COLUMN`.

    Returns
    -------
    StreamNodeDepletionReport
        Per-stream-node depletion results.

    Raises
    ------
    BudgetOutputMissingError
        Either model didn't declare a stream node budget output, or the
        declared file doesn't exist.
    KeyError
        ``sa_column`` not present in the budget headers.
    """
    from pyiwfm.io.budget import BudgetReader

    baseline_budget = _resolve_budget_file_from_model(
        baseline_model, kind="node", model_label="baseline"
    )
    scenario_budget = _resolve_budget_file_from_model(
        scenario_model, kind="node", model_label="scenario"
    )

    baseline_reader = BudgetReader(str(baseline_budget))
    scenario_reader = BudgetReader(str(scenario_budget))

    # Stream-node budget locations are stream nodes; the readers expose
    # them as a list of names like "1", "2", etc. (or "Node 1", etc.,
    # depending on IWFM version). We treat the location *index* + 1 as
    # the canonical 1-based stream node ID, mirroring the reach-level
    # convention. If a specific subset was requested, filter by that.
    locations = baseline_reader.locations
    if node_ids is not None:
        wanted = set(node_ids)
        location_indices = [i for i in range(len(locations)) if (i + 1) in wanted]
        # Surface obviously-bogus IDs early so the user gets a helpful error
        unknown = wanted - {i + 1 for i in range(len(locations))}
        if unknown:
            raise ValueError(
                f"node_ids {sorted(unknown)} are not present in the stream node "
                f"budget (it has {len(locations)} locations, ids 1..{len(locations)})"
            )
    else:
        location_indices = list(range(len(locations)))

    results: list[StreamNodeDepletionResult] = []

    for loc_idx in location_indices:
        base_times, base_values = baseline_reader.get_values(loc_idx)
        _scen_times, scen_values = scenario_reader.get_values(loc_idx)
        headers = baseline_reader.get_column_headers(loc_idx)

        base_flux = _extract_named_column(headers, base_values, sa_column)
        scen_flux = _extract_named_column(headers, scen_values, sa_column)

        n_ts = min(len(base_flux), len(scen_flux))
        base_flux = base_flux[:n_ts]
        scen_flux = scen_flux[:n_ts]

        depletion = base_flux - scen_flux
        cumulative = np.cumsum(depletion)
        times = _format_times(base_times[:n_ts], baseline_reader)
        max_dep_idx = int(np.argmax(np.abs(depletion))) if depletion.size else 0

        results.append(
            StreamNodeDepletionResult(
                stream_node_id=loc_idx + 1,
                times=times,
                baseline_sa_flux=base_flux,
                scenario_sa_flux=scen_flux,
                depletion=depletion,
                cumulative_depletion=cumulative,
                max_depletion=float(np.max(np.abs(depletion))) if depletion.size else 0.0,
                max_depletion_timestep=max_dep_idx,
                total_depletion=float(cumulative[-1]) if cumulative.size else 0.0,
            )
        )

    total_max = max((r.max_depletion for r in results), default=0.0)
    total_cum = sum(r.total_depletion for r in results)
    n_timesteps = results[0].depletion.shape[0] if results else 0

    return StreamNodeDepletionReport(
        results=results,
        n_stream_nodes=len(results),
        n_timesteps=n_timesteps,
        total_max_depletion=total_max,
        total_cumulative_depletion=total_cum,
    )


def _find_stream_budget(results_dir: Path, budget_type: str) -> Path:
    """Find stream budget HDF5 file in a results directory.

    Parameters
    ----------
    results_dir : Path
        Directory to search for stream budget files.
    budget_type : str
        Budget type name (unused in current implementation but reserved for
        future multi-budget disambiguation).

    Returns
    -------
    Path
        Path to the stream budget file.

    Raises
    ------
    FileNotFoundError
        If no stream budget file is found.
    """
    # Look for common stream budget file patterns
    patterns = [
        "*.hdf",
        "*.h5",
        "*Stream*Budget*.hdf5",
        "*Stream*Budget*.hdf",
        "*stream*budget*",
    ]

    for pattern in patterns:
        matches = list(results_dir.glob(pattern))
        for m in matches:
            if "stream" in m.stem.lower() and "budget" in m.stem.lower():
                return m

    # Fallback: try any HDF file
    for pattern in ["*.hdf", "*.h5", "*.hdf5"]:
        matches = list(results_dir.glob(pattern))
        if matches:
            return matches[0]

    msg = f"No stream budget file found in {results_dir}"
    raise FileNotFoundError(msg)


def _extract_named_column(
    headers: list[str],
    values: NDArray[np.float64],
    column_name: str,
) -> NDArray[np.float64]:
    """Extract a single column by **exact** header match.

    Header matching is exact (after stripping whitespace). The previous
    implementation did substring matching against ``"gain from gw"`` /
    ``"stream-aquifer"``, which silently misclassified columns when the
    budget file used a different naming scheme. Exact matching trades
    that fragility for a clear failure: if ``column_name`` isn't in
    ``headers``, this raises :class:`KeyError` with the available column
    list so the caller can pick the right one.

    Parameters
    ----------
    headers
        Column header names returned by ``BudgetReader.get_column_headers``.
    values
        Budget data, shape ``(n_timesteps, n_columns)``.
    column_name
        Exact column header to extract.

    Returns
    -------
    numpy.ndarray
        1-D array of values, length ``n_timesteps``.

    Raises
    ------
    KeyError
        If ``column_name`` isn't present in ``headers``. The message
        lists the available columns.
    """
    stripped = [h.strip() for h in headers]
    if column_name not in stripped:
        raise KeyError(
            f"Column {column_name!r} not found in budget headers. "
            f"Available columns: {stripped}. "
            f"If your model uses a non-standard column name, pass "
            f"``sa_column='<your column name>'``."
        )
    idx = stripped.index(column_name)
    result: NDArray[np.float64] = values[:, idx].copy()
    return result


# Backward-compat alias: some external callers may import the old name.
# It now requires an explicit ``column_name``; the substring-fallback
# behavior is gone.
def _extract_stream_flow(
    headers: list[str],
    values: NDArray[np.float64],
    column_name: str = DEFAULT_SA_COLUMN,
) -> NDArray[np.float64]:
    """Backward-compat shim — see :func:`_extract_named_column`.

    Pre-Phase-2.2.a-iv this function did substring matching against
    ``"gain from gw"`` / ``"stream-aquifer"``. It now requires an exact
    column name (defaulting to :data:`DEFAULT_SA_COLUMN`).
    """
    return _extract_named_column(headers, values, column_name)


def _resolve_budget_file_from_model(
    model: IWFMModel,
    kind: Literal["reach", "node"],
    *,
    model_label: str,
) -> Path:
    """Return the model-declared budget file path, validated.

    Looks up the path in :attr:`IWFMModel.metadata` (populated during
    load from the stream main file's ``STRMRCHBUDFL`` / ``STNDBUDFL``
    entries) and confirms the file exists on disk.

    Raises :class:`BudgetOutputMissingError` with an actionable message
    if the model didn't declare the output or the file is missing
    (typically because the simulation didn't run or didn't request the
    output).

    The ``model_label`` shows up in the error message
    (``"baseline"`` / ``"scenario"``) so the operator knows which model
    is misconfigured.
    """
    metadata_key = "stream_budget_file" if kind == "reach" else "stream_node_budget_file"
    raw = model.metadata.get(metadata_key)
    if not raw:
        raise BudgetOutputMissingError(
            model_label=model_label,
            kind=kind,
            metadata_key=metadata_key,
            reason="the stream main file did not declare this budget output",
        )
    path = Path(raw)
    if not path.exists():
        raise BudgetOutputMissingError(
            model_label=model_label,
            kind=kind,
            metadata_key=metadata_key,
            reason=f"declared file does not exist on disk: {path}",
        )
    return path


def _format_times(
    times_array: NDArray[np.float64],
    reader: object,
) -> list[str]:
    """Convert time array to string list.

    Parameters
    ----------
    times_array : NDArray[np.float64]
        Array of time values (Julian days or similar).
    reader : object
        Budget reader instance, checked for header/timestep metadata.

    Returns
    -------
    list[str]
        ISO-format date strings or integer indices as fallback.
    """
    try:
        start_dt = getattr(getattr(reader, "header", None), "timestep", None)
        if start_dt and hasattr(start_dt, "start_datetime") and start_dt.start_datetime:
            from datetime import timedelta

            base = start_dt.start_datetime
            return [(base + timedelta(days=float(t))).isoformat()[:10] for t in times_array]
    except Exception:  # noqa: BLE001
        pass
    return [str(int(t)) for t in times_array]


# ---------------------------------------------------------------------------
# Report writers (Phase 2.2.a-i)
#
# These export a ``StreamDepletionReport`` to stakeholder-friendly formats.
# They are *report writers*, not model-input writers — the depletion analysis
# is a derivation, not IWFM input. See ``docs/user_guide/inputs_vs_outputs``.
# ---------------------------------------------------------------------------


# Column order is part of the public CSV contract. Downstream consumers
# (plotting tools, dashboards) depend on this ordering, so changes need a
# CHANGELOG note.
_CSV_COLUMNS: tuple[str, ...] = (
    "reach_id",
    "reach_name",
    "time",
    "baseline_flow",
    "scenario_flow",
    "depletion",
    "cumulative_depletion",
)


def write_stream_depletion_csv(
    report: StreamDepletionReport,
    output_path: str | Path,
) -> Path:
    """Write a long-format CSV: one row per reach × timestep.

    Columns: ``reach_id, reach_name, time, baseline_flow, scenario_flow,
    depletion, cumulative_depletion``. Suitable for plotting in Excel,
    pandas, or Tableau.

    Parameters
    ----------
    report
        The :class:`StreamDepletionReport` to write.
    output_path
        Destination CSV path. Parent directories must exist.

    Returns
    -------
    pathlib.Path
        The path written.
    """
    import csv

    output_path = Path(output_path)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(_CSV_COLUMNS)
        for r in report.results:
            for i, t in enumerate(r.times):
                writer.writerow(
                    [
                        r.reach_id,
                        r.reach_name,
                        t,
                        f"{float(r.baseline_flow[i]):.6g}",
                        f"{float(r.scenario_flow[i]):.6g}",
                        f"{float(r.depletion[i]):.6g}",
                        f"{float(r.cumulative_depletion[i]):.6g}",
                    ]
                )
    return output_path


def write_stream_depletion_json(
    report: StreamDepletionReport,
    output_path: str | Path,
    *,
    indent: int | None = 2,
) -> Path:
    """Write the full report as JSON via :meth:`StreamDepletionReport.to_dict`.

    Suitable for downstream Python or JS consumers that want the structured
    representation without recomputing depletion.
    """
    import json

    output_path = Path(output_path)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(report.to_dict(), f, indent=indent)
    return output_path


def _safe_sheet_name(name: str, fallback: str) -> str:
    """Return an Excel-safe sheet name (≤31 chars, no invalid chars)."""
    invalid = r"[]:*?/\\"
    clean = name or fallback
    for ch in invalid:
        clean = clean.replace(ch, "_")
    clean = clean.strip()
    if not clean:
        clean = fallback
    return clean[:31]


def write_stream_depletion_excel(
    report: StreamDepletionReport,
    output_path: str | Path,
) -> Path:
    """Write a multi-sheet xlsx report.

    Layout:

    - **Summary** — one row per reach with ``reach_id``, ``reach_name``,
      ``max_depletion``, ``max_depletion_timestep``, ``total_depletion``,
      plus a totals row.
    - **One sheet per reach** — long-format time series matching the CSV
      columns (without the ``reach_id`` / ``reach_name`` prefix).

    Mirrors the openpyxl-based pattern in :mod:`pyiwfm.io.budget.excel`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    output_path = Path(output_path)
    wb = Workbook()
    bold = Font(bold=True)

    # Summary sheet
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(
        [
            "reach_id",
            "reach_name",
            "max_depletion",
            "max_depletion_timestep",
            "total_depletion",
        ]
    )
    for cell in summary[1]:
        cell.font = bold
    for r in report.results:
        summary.append(
            [
                r.reach_id,
                r.reach_name,
                round(float(r.max_depletion), 6),
                int(r.max_depletion_timestep),
                round(float(r.total_depletion), 6),
            ]
        )
    summary.append([])
    summary.append(
        [
            "TOTAL",
            "",
            round(float(report.total_max_depletion), 6),
            "",
            round(float(report.total_cumulative_depletion), 6),
        ]
    )

    # One sheet per reach
    used_names: set[str] = {"Summary"}
    for r in report.results:
        base = _safe_sheet_name(r.reach_name, fallback=f"Reach_{r.reach_id}")
        # Disambiguate duplicates
        name = base
        suffix = 2
        while name in used_names:
            tail = f"_{suffix}"
            name = (base[: 31 - len(tail)] + tail) if len(base) + len(tail) > 31 else base + tail
            suffix += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        ws.append(
            [
                "time",
                "baseline_flow",
                "scenario_flow",
                "depletion",
                "cumulative_depletion",
            ]
        )
        for cell in ws[1]:
            cell.font = bold
        for i, t in enumerate(r.times):
            ws.append(
                [
                    t,
                    round(float(r.baseline_flow[i]), 6),
                    round(float(r.scenario_flow[i]), 6),
                    round(float(r.depletion[i]), 6),
                    round(float(r.cumulative_depletion[i]), 6),
                ]
            )

    wb.save(str(output_path))
    return output_path


# Hook the dispatcher onto StreamDepletionReport. Kept at module scope (rather
# than in the @dataclass body) so the writers are forward references — the
# dataclass declaration above stays narrow and the writers can reference the
# class.
def _report_write(
    self: StreamDepletionReport,
    path: str | Path,
    format: str | None = None,
) -> Path:
    """Dispatch to a CSV / JSON / Excel writer by file extension or by
    explicit ``format``.

    Parameters
    ----------
    path
        Destination path.
    format
        Optional explicit format: ``"csv"``, ``"json"``, ``"xlsx"``.
        When ``None`` (default), the format is chosen from the file
        extension.

    Raises
    ------
    ValueError
        If ``format`` is unknown, or if it is ``None`` and the file
        extension isn't recognized.
    """
    path = Path(path)
    fmt = (format or path.suffix.lstrip(".")).lower()
    if fmt == "csv":
        return write_stream_depletion_csv(self, path)
    if fmt == "json":
        return write_stream_depletion_json(self, path)
    if fmt in ("xlsx", "excel"):
        return write_stream_depletion_excel(self, path)
    raise ValueError(
        f"Unknown format {format!r} (extension {path.suffix!r}); expected one of: csv, json, xlsx"
    )


StreamDepletionReport.write = _report_write  # type: ignore[method-assign,attr-defined]
