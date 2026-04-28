"""Drawdown computation from IWFM head data.

**Read-only analysis.** This module computes drawdown by comparing
heads from two timesteps of one model run; the result is a *report*,
not a model input. Tabular and graphical exports for stakeholders are
documented in ``docs/user_guide/inputs_vs_outputs.rst``.

The two report shapes:

- :class:`DrawdownTimeSeriesReport` — drawdown vs time at a small set
  of (node, layer) locations (e.g. observation wells). Used by
  :func:`write_drawdown_timeseries_csv` /
  :func:`write_drawdown_timeseries_excel` and the time-series plots in
  :mod:`pyiwfm.visualization.plot_drawdown`.
- :class:`DrawdownSnapshot` — drawdown at every node at a single
  timestep (or the per-node max across all timesteps). Used by the
  cone-of-depression and max-drawdown maps in
  :mod:`pyiwfm.visualization.map_drawdown`.

This module provides the :class:`DrawdownComputer` class for computing
drawdown (head change) relative to a reference timestep.  Drawdown is
defined as ``head(t_ref) - head(t)`` so that positive values indicate
water level decline.

Sentinel values (< -9000) used by IWFM to represent dry cells are
treated as NaN.
"""

from __future__ import annotations

import dataclasses
import logging
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np
from numpy.typing import NDArray

if TYPE_CHECKING:
    from collections.abc import Sequence

    from pyiwfm.core.mesh import AppGrid
    from pyiwfm.io.timeseries_io import LazyNodalLoader

logger = logging.getLogger(__name__)

_SENTINEL = -9000.0


class DrawdownComputer:
    """Compute drawdown (head change) relative to a reference timestep.

    Drawdown = head(t_ref) - head(t)  (positive = water level decline)

    Parameters
    ----------
    head_loader : LazyNodalLoader
        Loader providing lazy access to per-timestep head arrays.
    """

    def __init__(self, head_loader: LazyNodalLoader) -> None:
        self._loader = head_loader

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_layer_values(self, timestep: int, layer: int) -> NDArray[np.float64]:
        """Return head values for a single layer, with sentinels as NaN.

        Parameters
        ----------
        timestep : int
            0-based timestep index.
        layer : int
            1-based layer number.

        Returns
        -------
        NDArray[np.float64]
            Shape ``(n_nodes,)``.  Dry-cell sentinels replaced with NaN.
        """
        frame = self._loader.get_frame(timestep)
        layer_idx = layer - 1
        if layer_idx >= frame.shape[1]:
            raise IndexError(f"Layer {layer} out of range [1, {frame.shape[1]}]")
        col: NDArray[np.float64] = frame[:, layer_idx].copy()
        col[col < _SENTINEL] = np.nan
        return col

    def _validate_timestep(self, timestep: int) -> None:
        if timestep < 0 or timestep >= self._loader.n_frames:
            raise IndexError(f"Timestep {timestep} out of range [0, {self._loader.n_frames})")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def compute_drawdown(
        self,
        timestep: int,
        reference_timestep: int = 0,
        layer: int = 1,
    ) -> NDArray[np.float64]:
        """Compute per-node drawdown for a single timestep vs reference.

        Parameters
        ----------
        timestep : int
            0-based target timestep index.
        reference_timestep : int
            0-based reference timestep index.  Default is the first
            timestep (0).
        layer : int
            1-based layer number.

        Returns
        -------
        NDArray[np.float64]
            Array of shape ``(n_nodes,)``.  Positive values mean
            water level decline.  NaN where either timestep has a
            dry-cell sentinel.
        """
        self._validate_timestep(timestep)
        self._validate_timestep(reference_timestep)

        ref = self._get_layer_values(reference_timestep, layer)
        cur = self._get_layer_values(timestep, layer)

        drawdown: NDArray[np.float64] = ref - cur
        return drawdown

    def compute_drawdown_by_element(
        self,
        timestep: int,
        reference_timestep: int = 0,
        layer: int = 1,
        grid: AppGrid | None = None,
    ) -> NDArray[np.float64]:
        """Compute per-element drawdown (average of element node values).

        Each element's drawdown is the arithmetic mean of its vertex
        node drawdown values.

        Parameters
        ----------
        timestep : int
            0-based target timestep index.
        reference_timestep : int
            0-based reference timestep index.
        layer : int
            1-based layer number.
        grid : AppGrid or None
            The model grid.  Required to look up element vertices.

        Returns
        -------
        NDArray[np.float64]
            Array of shape ``(n_elements,)`` ordered by element ID.
            NaN where any contributing node is NaN.

        Raises
        ------
        ValueError
            If *grid* is ``None``.
        """
        if grid is None:
            raise ValueError("grid is required for element-level drawdown")

        node_dd = self.compute_drawdown(timestep, reference_timestep, layer)

        # Sort elements by ID for deterministic output order
        sorted_elems = sorted(grid.elements.values(), key=lambda e: e.id)
        n_elems = len(sorted_elems)
        result = np.empty(n_elems, dtype=np.float64)

        for i, elem in enumerate(sorted_elems):
            # vertices are 1-based node IDs
            indices = np.array([nid - 1 for nid in elem.vertices])
            vals = node_dd[indices]
            result[i] = float(np.nanmean(vals))

        return result

    def compute_drawdown_range(
        self,
        reference_timestep: int = 0,
        layer: int = 1,
        max_frames: int = 0,
    ) -> tuple[float, float]:
        """Compute robust min/max drawdown across all timesteps.

        Uses the 2nd and 98th percentile to exclude outliers.

        Parameters
        ----------
        reference_timestep : int
            0-based reference timestep index.
        layer : int
            1-based layer number.
        max_frames : int
            If > 0, sample at most this many evenly-spaced frames
            instead of scanning every timestep.

        Returns
        -------
        tuple[float, float]
            ``(min_drawdown, max_drawdown)`` across all (sampled)
            timesteps.  Falls back to ``(0.0, 1.0)`` if no valid
            data exists.
        """
        self._validate_timestep(reference_timestep)

        total = self._loader.n_frames
        if total == 0:
            return (0.0, 1.0)

        # Determine which frames to sample
        if max_frames > 0 and max_frames < total:
            indices = np.unique(np.linspace(0, total - 1, max_frames, dtype=int))
        else:
            indices = np.arange(total)

        ref = self._get_layer_values(reference_timestep, layer)
        all_valid: list[float] = []

        for idx in indices:
            cur = self._get_layer_values(int(idx), layer)
            dd = ref - cur
            valid = dd[np.isfinite(dd)]
            if len(valid) > 0:
                all_valid.extend(valid.tolist())

        if not all_valid:
            return (0.0, 1.0)

        arr = np.array(all_valid)
        lo = float(np.percentile(arr, 2.0))
        hi = float(np.percentile(arr, 98.0))
        return (round(lo, 3), round(hi, 3))

    def compute_max_drawdown_map(
        self,
        reference_timestep: int = 0,
        layer: int = 1,
    ) -> NDArray[np.float64]:
        """Compute maximum drawdown at each node across all timesteps.

        Parameters
        ----------
        reference_timestep : int
            0-based reference timestep index.
        layer : int
            1-based layer number.

        Returns
        -------
        NDArray[np.float64]
            Array of shape ``(n_nodes,)``.  Each entry is the maximum
            drawdown observed at that node over all timesteps.  NaN
            where the node is always dry.
        """
        self._validate_timestep(reference_timestep)

        ref = self._get_layer_values(reference_timestep, layer)
        n_nodes = len(ref)
        max_dd = np.full(n_nodes, np.nan, dtype=np.float64)

        for ts in range(self._loader.n_frames):
            cur = self._get_layer_values(ts, layer)
            dd = ref - cur
            max_dd = np.fmax(max_dd, dd)

        return max_dd

    # ------------------------------------------------------------------
    # Report builders (Phase 2 drawdown)
    # ------------------------------------------------------------------

    def build_timeseries_report(
        self,
        locations: Sequence[tuple[int, int]],
        reference_timestep: int = 0,
    ) -> DrawdownTimeSeriesReport:
        """Build a per-location time-series drawdown report.

        Useful for plotting drawdown over time at a small set of
        observation wells (or any chosen ``(node_id, layer)`` pairs).
        For each location, walks every timestep once, so cost is
        O(n_locations × n_timesteps × n_nodes) — keep ``locations``
        small (tens to low hundreds, not thousands).

        Parameters
        ----------
        locations
            Iterable of ``(node_id, layer)`` pairs (both 1-based, IWFM
            convention).
        reference_timestep
            0-based reference timestep. Drawdown is computed as
            ``head(reference_timestep) - head(t)`` so positive values
            mean decline.

        Returns
        -------
        DrawdownTimeSeriesReport

        Raises
        ------
        IndexError
            ``reference_timestep`` out of range, or any
            ``(node_id, layer)`` pair refers to a node/layer outside
            the loader's bounds.
        """
        self._validate_timestep(reference_timestep)
        loc_list = list(locations)
        if not loc_list:
            return DrawdownTimeSeriesReport(
                locations=[],
                n_locations=0,
                n_timesteps=self._loader.n_frames,
                reference_timestep=reference_timestep,
                times=_format_loader_times(self._loader),
            )

        n_layers = self._loader.n_layers
        n_nodes = self._loader.n_nodes
        n_ts = self._loader.n_frames
        for node_id, layer in loc_list:
            if not 1 <= node_id <= n_nodes:
                raise IndexError(f"node_id {node_id} out of range [1, {n_nodes}]")
            if not 1 <= layer <= n_layers:
                raise IndexError(f"layer {layer} out of range [1, {n_layers}]")

        # Reference values at each (node, layer): cache by layer to avoid
        # re-reading the reference frame for every location.
        ref_by_layer: dict[int, NDArray[np.float64]] = {}
        for _, layer in loc_list:
            if layer not in ref_by_layer:
                ref_by_layer[layer] = self._get_layer_values(reference_timestep, layer)

        # Allocate per-location time series, then walk each timestep
        # once to fill them in. This trades an extra (n_locations,
        # n_timesteps) array for one frame load per timestep instead of
        # n_locations × n_timesteps frame loads.
        per_loc: list[NDArray[np.float64]] = [np.empty(n_ts, dtype=np.float64) for _ in loc_list]
        for ts in range(n_ts):
            # One frame read per (ts, layer); cache within the timestep
            cur_by_layer: dict[int, NDArray[np.float64]] = {}
            for i, (node_id, layer) in enumerate(loc_list):
                if layer not in cur_by_layer:
                    cur_by_layer[layer] = self._get_layer_values(ts, layer)
                cur = cur_by_layer[layer]
                per_loc[i][ts] = ref_by_layer[layer][node_id - 1] - cur[node_id - 1]

        times = _format_loader_times(self._loader)
        results = []
        for (node_id, layer), dd in zip(loc_list, per_loc, strict=True):
            finite = dd[np.isfinite(dd)]
            if finite.size:
                max_idx = int(np.nanargmax(np.abs(dd)))
                max_val = float(dd[max_idx])
                final_val = float(dd[-1]) if np.isfinite(dd[-1]) else float("nan")
            else:
                max_idx = 0
                max_val = float("nan")
                final_val = float("nan")
            results.append(
                DrawdownAtLocation(
                    node_id=node_id,
                    layer=layer,
                    times=times,
                    drawdown=dd,
                    max_drawdown=max_val,
                    max_drawdown_timestep=max_idx,
                    final_drawdown=final_val,
                )
            )

        return DrawdownTimeSeriesReport(
            locations=results,
            n_locations=len(results),
            n_timesteps=n_ts,
            reference_timestep=reference_timestep,
            times=times,
        )

    def build_snapshot(
        self,
        timestep: int,
        layer: int = 1,
        reference_timestep: int = 0,
    ) -> DrawdownSnapshot:
        """Build a per-node drawdown snapshot at a single timestep.

        This is the cone-of-depression view: drawdown at every node at
        a chosen timestep, suitable for a colored map of the spatial
        extent of drawdown.

        Parameters
        ----------
        timestep
            0-based target timestep.
        layer
            1-based layer.
        reference_timestep
            0-based reference timestep. Default 0.

        Returns
        -------
        DrawdownSnapshot
        """
        drawdown = self.compute_drawdown(timestep, reference_timestep, layer)
        node_ids = np.arange(1, len(drawdown) + 1, dtype=np.int32)
        return DrawdownSnapshot(
            timestep=timestep,
            layer=layer,
            reference_timestep=reference_timestep,
            kind="single",
            time_label=_format_timestep(self._loader, timestep),
            node_ids=node_ids,
            drawdown=drawdown,
            n_nodes=len(drawdown),
        )

    def build_max_snapshot(
        self,
        layer: int = 1,
        reference_timestep: int = 0,
    ) -> DrawdownSnapshot:
        """Build a per-node max-across-time drawdown snapshot.

        For each node, returns the maximum drawdown observed across all
        timesteps. Useful for showing worst-case impact of a stress
        period regardless of when each node sees its peak decline.

        Parameters
        ----------
        layer
            1-based layer.
        reference_timestep
            0-based reference timestep.

        Returns
        -------
        DrawdownSnapshot
            With ``kind="max"`` and ``timestep=-1`` (sentinel meaning
            "across all timesteps").
        """
        drawdown = self.compute_max_drawdown_map(reference_timestep, layer)
        node_ids = np.arange(1, len(drawdown) + 1, dtype=np.int32)
        return DrawdownSnapshot(
            timestep=-1,
            layer=layer,
            reference_timestep=reference_timestep,
            kind="max",
            time_label=f"max over {self._loader.n_frames} timesteps",
            node_ids=node_ids,
            drawdown=drawdown,
            n_nodes=len(drawdown),
        )


# ---------------------------------------------------------------------------
# Report dataclasses
# ---------------------------------------------------------------------------


def _format_loader_times(loader: LazyNodalLoader) -> list[str]:
    """Convert the loader's ``datetime`` timestamps to ISO date strings."""
    times = loader.times
    return [t.isoformat()[:19] if hasattr(t, "isoformat") else str(t) for t in times]


def _format_timestep(loader: LazyNodalLoader, timestep: int) -> str:
    """Format a single timestep index as an ISO date string when possible."""
    if 0 <= timestep < len(loader.times):
        t = loader.times[timestep]
        return t.isoformat()[:19] if hasattr(t, "isoformat") else str(t)
    return f"timestep {timestep}"


@dataclasses.dataclass
class DrawdownAtLocation:
    """Drawdown over time at a single (node, layer) location.

    Attributes
    ----------
    node_id
        1-based node ID (IWFM convention).
    layer
        1-based aquifer layer.
    times
        Timestamp strings, one per element of ``drawdown``.
    drawdown
        Per-timestep drawdown values (positive = water-level decline).
        NaN where the head is a dry-cell sentinel at that timestep.
    max_drawdown
        Largest absolute drawdown across all finite timesteps. ``NaN``
        if every timestep is dry.
    max_drawdown_timestep
        0-based timestep index where ``max_drawdown`` occurs.
    final_drawdown
        Drawdown at the last timestep (NaN if dry).
    """

    node_id: int
    layer: int
    times: list[str]
    drawdown: NDArray[np.float64]
    max_drawdown: float
    max_drawdown_timestep: int
    final_drawdown: float

    def to_dict(self) -> dict[str, object]:
        """Serialize for API responses / JSON export."""
        return {
            "node_id": self.node_id,
            "layer": self.layer,
            "times": self.times,
            "drawdown": [None if not np.isfinite(v) else round(float(v), 4) for v in self.drawdown],
            "max_drawdown": (
                None if not np.isfinite(self.max_drawdown) else round(float(self.max_drawdown), 4)
            ),
            "max_drawdown_timestep": int(self.max_drawdown_timestep),
            "final_drawdown": (
                None
                if not np.isfinite(self.final_drawdown)
                else round(float(self.final_drawdown), 4)
            ),
        }


@dataclasses.dataclass
class DrawdownTimeSeriesReport:
    """Time-series drawdown at a set of (node, layer) locations.

    Suitable for tabular export (CSV/Excel) and per-location plots.
    Build via :meth:`DrawdownComputer.build_timeseries_report`.

    Attributes
    ----------
    locations
        One :class:`DrawdownAtLocation` per (node, layer) requested.
    n_locations
        ``len(locations)``.
    n_timesteps
        Total timesteps in the underlying loader.
    reference_timestep
        0-based timestep that drawdown was computed against.
    times
        Timestamp strings, shared across all locations.
    """

    locations: list[DrawdownAtLocation]
    n_locations: int
    n_timesteps: int
    reference_timestep: int
    times: list[str]

    def to_dict(self) -> dict[str, object]:
        """Serialize for API responses / JSON export."""
        return {
            "n_locations": self.n_locations,
            "n_timesteps": self.n_timesteps,
            "reference_timestep": self.reference_timestep,
            "times": self.times,
            "locations": [loc.to_dict() for loc in self.locations],
        }


@dataclasses.dataclass
class DrawdownSnapshot:
    """Drawdown at every node at a single timestep (or per-node max).

    Suitable for cone-of-depression maps and GeoJSON export. Build via
    :meth:`DrawdownComputer.build_snapshot` (single timestep) or
    :meth:`DrawdownComputer.build_max_snapshot` (max across all
    timesteps).

    Attributes
    ----------
    timestep
        0-based timestep index, or ``-1`` for ``kind="max"``.
    layer
        1-based aquifer layer.
    reference_timestep
        0-based reference timestep.
    kind
        ``"single"`` (snapshot at one timestep) or ``"max"`` (per-node
        max across all timesteps).
    time_label
        Human-readable timestamp / description for plot titles.
    node_ids
        1-based node IDs, length ``n_nodes``.
    drawdown
        Per-node drawdown values (positive = decline, NaN for dry).
    n_nodes
        ``len(drawdown)``.
    """

    timestep: int
    layer: int
    reference_timestep: int
    kind: str  # "single" or "max"
    time_label: str
    node_ids: NDArray[np.int32]
    drawdown: NDArray[np.float64]
    n_nodes: int

    def to_dict(self) -> dict[str, object]:
        """Serialize for API responses / JSON export.

        Drawdown values are listed in ``node_ids`` order. NaN dry-cell
        values are emitted as ``null``.
        """
        return {
            "timestep": int(self.timestep),
            "layer": int(self.layer),
            "reference_timestep": int(self.reference_timestep),
            "kind": self.kind,
            "time_label": self.time_label,
            "n_nodes": int(self.n_nodes),
            "node_ids": [int(n) for n in self.node_ids],
            "drawdown": [None if not np.isfinite(v) else round(float(v), 4) for v in self.drawdown],
        }


# ---------------------------------------------------------------------------
# Tabular report writers
# ---------------------------------------------------------------------------


# Column order is part of the public CSV contract.
_DRAWDOWN_TS_CSV_COLUMNS: tuple[str, ...] = (
    "node_id",
    "layer",
    "time",
    "drawdown",
)


def write_drawdown_timeseries_csv(
    report: DrawdownTimeSeriesReport,
    output_path: str | Path,
) -> Path:
    """Write a long-format CSV: one row per location × timestep.

    Columns: ``node_id, layer, time, drawdown``. Suitable for plotting
    in Excel, pandas, or any spreadsheet tool. NaN dry-cell values are
    written as the empty string.
    """
    import csv

    output_path = Path(output_path)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(_DRAWDOWN_TS_CSV_COLUMNS)
        for loc in report.locations:
            for i, t in enumerate(loc.times):
                v = loc.drawdown[i]
                writer.writerow(
                    [
                        loc.node_id,
                        loc.layer,
                        t,
                        f"{float(v):.6g}" if np.isfinite(v) else "",
                    ]
                )
    return output_path


def write_drawdown_timeseries_json(
    report: DrawdownTimeSeriesReport,
    output_path: str | Path,
    *,
    indent: int | None = 2,
) -> Path:
    """Write the full report as JSON via :meth:`DrawdownTimeSeriesReport.to_dict`."""
    import json

    output_path = Path(output_path)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(report.to_dict(), f, indent=indent)
    return output_path


def _safe_drawdown_sheet_name(loc: DrawdownAtLocation, fallback_idx: int) -> str:
    """Build an Excel-safe sheet name for a (node, layer) location."""
    name = f"Node{loc.node_id}_L{loc.layer}"
    invalid = r"[]:*?/\\"
    for ch in invalid:
        name = name.replace(ch, "_")
    name = name.strip()
    if not name:
        name = f"Loc_{fallback_idx + 1}"
    return name[:31]


def write_drawdown_timeseries_excel(
    report: DrawdownTimeSeriesReport,
    output_path: str | Path,
) -> Path:
    """Write a multi-sheet xlsx report.

    Layout:

    - **Summary** — one row per location with ``node_id``, ``layer``,
      ``max_drawdown``, ``max_drawdown_timestep``, ``final_drawdown``.
    - **One sheet per location** — long-format ``time, drawdown``.

    Mirrors the openpyxl-based pattern in
    :mod:`pyiwfm.io.stream_depletion`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    output_path = Path(output_path)
    wb = Workbook()
    bold = Font(bold=True)

    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(
        [
            "node_id",
            "layer",
            "max_drawdown",
            "max_drawdown_timestep",
            "final_drawdown",
        ]
    )
    for cell in summary[1]:
        cell.font = bold
    for loc in report.locations:
        summary.append(
            [
                loc.node_id,
                loc.layer,
                None if not np.isfinite(loc.max_drawdown) else round(float(loc.max_drawdown), 6),
                int(loc.max_drawdown_timestep),
                (
                    None
                    if not np.isfinite(loc.final_drawdown)
                    else round(float(loc.final_drawdown), 6)
                ),
            ]
        )

    used_names: set[str] = {"Summary"}
    for i, loc in enumerate(report.locations):
        base = _safe_drawdown_sheet_name(loc, i)
        name = base
        suffix = 2
        while name in used_names:
            tail = f"_{suffix}"
            name = (base[: 31 - len(tail)] + tail) if len(base) + len(tail) > 31 else base + tail
            suffix += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        ws.append(["time", "drawdown"])
        for cell in ws[1]:
            cell.font = bold
        for j, t in enumerate(loc.times):
            v = loc.drawdown[j]
            ws.append([t, None if not np.isfinite(v) else round(float(v), 6)])

    wb.save(str(output_path))
    return output_path


def _drawdown_report_write(
    self: DrawdownTimeSeriesReport,
    path: str | Path,
    format: str | None = None,
) -> Path:
    """Dispatch CSV / JSON / xlsx writer by extension or explicit ``format``.

    Parameters
    ----------
    path
        Destination path.
    format
        ``"csv"``, ``"json"``, or ``"xlsx"`` / ``"excel"``. ``None``
        (default) infers from the extension.

    Raises
    ------
    ValueError
        Unknown format.
    """
    path = Path(path)
    fmt = (format or path.suffix.lstrip(".")).lower()
    if fmt == "csv":
        return write_drawdown_timeseries_csv(self, path)
    if fmt == "json":
        return write_drawdown_timeseries_json(self, path)
    if fmt in ("xlsx", "excel"):
        return write_drawdown_timeseries_excel(self, path)
    raise ValueError(
        f"Unknown format {format!r} (extension {path.suffix!r}); expected one of: csv, json, xlsx"
    )


DrawdownTimeSeriesReport.write = _drawdown_report_write  # type: ignore[method-assign,attr-defined]
