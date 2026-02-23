"""
Budget Excel export for IWFM post-processing.

Generates Excel workbooks from IWFM budget HDF5 data with one sheet per
location, title lines, bold headers, and auto-fitted column widths —
matching the reference implementation in ``SGMOModeling/pywfm-budget-excel``.
"""

from __future__ import annotations

import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from pyiwfm.io.budget import BudgetReader
from pyiwfm.io.budget_control import BudgetControlConfig
from pyiwfm.io.budget_utils import filter_time_range, format_title_lines

logger = logging.getLogger(__name__)


def _make_sheet_name(name: str) -> str:
    """Create a valid Excel sheet name (max 31 chars, no invalid chars)."""
    invalid = r"[]:*?/\\"
    clean = name
    for ch in invalid:
        clean = clean.replace(ch, "_")
    return clean[:31]


def budget_to_excel(
    reader: BudgetReader,
    output_path: Path | str,
    length_factor: float = 1.0,
    area_factor: float = 1.0,
    volume_factor: float = 1.0,
    length_unit: str = "FEET",
    area_unit: str = "AC",
    volume_unit: str = "AC.FT.",
    location_ids: list[int] | None = None,
    begin_date: str | None = None,
    end_date: str | None = None,
    output_interval: str | None = None,
) -> Path:
    """Export budget data to an Excel workbook.

    One sheet per location.  Each sheet contains:

    - Title lines (bold) from ``ASCIIOutputInfo`` with unit substitution.
    - Column headers (bold).
    - Data rows with unit conversion applied.
    - Auto-fitted column widths.

    Parameters
    ----------
    reader : BudgetReader
        Open budget reader.
    output_path : Path or str
        Destination ``.xlsx`` file.
    length_factor, area_factor, volume_factor : float
        FACTLTOU / FACTAROU / FACTVLOU from the control file.
    length_unit, area_unit, volume_unit : str
        UNITLTOU / UNITAROU / UNITVLOU for title-line substitution.
    location_ids : list[int] or None
        1-based location IDs to include.  ``None`` or ``[-1]`` → all.
    begin_date, end_date : str or None
        IWFM datetime strings for time filtering.
    output_interval : str or None
        Not yet implemented — reserved for future resampling.

    Returns
    -------
    Path
        The written workbook path.
    """
    output_path = Path(output_path)

    # Determine which locations to export
    if location_ids is None or location_ids == [-1]:
        loc_indices = list(range(reader.n_locations))
    elif not location_ids:
        loc_indices = []
    else:
        loc_indices = [lid - 1 for lid in location_ids]

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    if wb.active is not None:
        wb.remove(wb.active)

    bold_font = Font(bold=True)

    for loc_idx in loc_indices:
        if loc_idx < 0 or loc_idx >= reader.n_locations:
            logger.warning("Skipping out-of-range location index %d", loc_idx + 1)
            continue

        loc_name = reader.locations[loc_idx]
        sheet_name = _make_sheet_name(loc_name)
        ws = wb.create_sheet(title=sheet_name)

        # --- Title lines ---
        titles = reader.header.ascii_output.titles
        area_val = (
            float(reader.header.areas[loc_idx])
            if reader.header.areas is not None and loc_idx < len(reader.header.areas)
            else None
        )
        formatted_titles = format_title_lines(
            titles,
            location_name=loc_name,
            area=area_val,
            length_unit=length_unit,
            area_unit=area_unit,
            volume_unit=volume_unit,
        )

        row_num = 1
        for title in formatted_titles:
            ws.cell(row=row_num, column=1, value=title).font = bold_font
            row_num += 1

        # Blank row between titles and headers
        row_num += 1

        # --- Column headers ---
        col_headers = reader.get_column_headers(loc_idx)
        time_header = "Time"
        all_headers = [time_header] + col_headers

        for col_idx_h, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx_h, value=header)
            cell.font = bold_font
        row_num += 1

        # --- Data rows (with unit conversion) ---
        df = reader.get_dataframe(
            loc_idx,
            length_factor=length_factor,
            area_factor=area_factor,
            volume_factor=volume_factor,
        )
        df = filter_time_range(df, begin_date, end_date)

        if isinstance(df.index, pd.DatetimeIndex):
            time_strings = [dt.strftime("%m/%d/%Y_%H:%M") for dt in df.index]
        else:
            time_strings = [str(t) for t in df.index]

        values = df.to_numpy(dtype=np.float64, na_value=np.nan)
        for i in range(len(df)):
            ws.cell(row=row_num, column=1, value=time_strings[i])
            for j in range(df.shape[1]):
                num = values[i, j]
                ws.cell(
                    row=row_num,
                    column=j + 2,
                    value=round(float(num), 4) if not np.isnan(num) else None,
                )
            row_num += 1

        # --- Auto-fit column widths ---
        for col_idx_w in range(1, len(all_headers) + 1):
            col_letter = get_column_letter(col_idx_w)
            max_width = 0
            for row_cells in ws.iter_rows(
                min_col=col_idx_w, max_col=col_idx_w, min_row=1, max_row=row_num - 1
            ):
                for cell in row_cells:
                    if cell.value is not None:
                        max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    # If no sheets were created, add a placeholder
    if not wb.sheetnames:
        wb.create_sheet(title="(no data)")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def budget_control_to_excel(config: BudgetControlConfig) -> list[Path]:
    """Process a full budget control file, generating one ``.xlsx`` per spec.

    Parameters
    ----------
    config : BudgetControlConfig
        Parsed budget control configuration.

    Returns
    -------
    list[Path]
        Paths of created workbook files.
    """
    created: list[Path] = []

    for spec in config.budgets:
        if not spec.hdf_file.exists():
            logger.warning("Budget HDF5 file not found: %s", spec.hdf_file)
            continue

        reader = BudgetReader(spec.hdf_file)

        # Ensure output has .xlsx extension
        out = spec.output_file
        if out.suffix.lower() not in (".xlsx", ".xls"):
            out = out.with_suffix(".xlsx")

        result = budget_to_excel(
            reader=reader,
            output_path=out,
            length_factor=config.length_factor,
            area_factor=config.area_factor,
            volume_factor=config.volume_factor,
            length_unit=config.length_unit,
            area_unit=config.area_unit,
            volume_unit=config.volume_unit,
            location_ids=spec.location_ids if spec.location_ids else None,
            begin_date=config.begin_date,
            end_date=config.end_date,
            output_interval=spec.output_interval,
        )
        created.append(result)
        logger.info("Wrote %s (%d locations)", result, reader.n_locations)

    return created
