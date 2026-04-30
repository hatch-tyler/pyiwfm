"""Tests for the stream depletion report writers (Phase 2.2.a-i).

Covers:

- ``write_stream_depletion_csv`` — long-format CSV with golden header
- ``write_stream_depletion_json`` — JSON wrapping ``to_dict()``
- ``write_stream_depletion_excel`` — multi-sheet xlsx (Summary + per-reach)
- ``StreamDepletionReport.write(path, format=None)`` — extension dispatch
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from pyiwfm.io.streams.depletion import (
    StreamDepletionReport,
    StreamDepletionResult,
    write_stream_depletion_csv,
    write_stream_depletion_excel,
    write_stream_depletion_json,
)


def _make_report(n_reaches: int = 2, n_timesteps: int = 5) -> StreamDepletionReport:
    """Build a small fixture report with deterministic values."""
    rng = np.random.default_rng(0)
    times = [f"2024-{m:02d}-01" for m in range(1, n_timesteps + 1)]
    results = []
    for ri in range(n_reaches):
        base = rng.uniform(50.0, 100.0, n_timesteps)
        scen = base - rng.uniform(0.5, 5.0, n_timesteps)
        depletion = base - scen
        cumulative = np.cumsum(depletion)
        results.append(
            StreamDepletionResult(
                reach_id=ri + 1,
                reach_name=f"Reach {ri + 1}",
                times=times,
                baseline_flow=base,
                scenario_flow=scen,
                depletion=depletion,
                cumulative_depletion=cumulative,
                max_depletion=float(np.max(np.abs(depletion))),
                max_depletion_timestep=int(np.argmax(np.abs(depletion))),
                total_depletion=float(cumulative[-1]),
            )
        )
    return StreamDepletionReport(
        results=results,
        n_reaches=n_reaches,
        n_timesteps=n_timesteps,
        total_max_depletion=max(r.max_depletion for r in results),
        total_cumulative_depletion=sum(r.total_depletion for r in results),
    )


class TestCsvWriter:
    def test_writes_expected_columns(self, tmp_path: Path):
        report = _make_report(n_reaches=2, n_timesteps=3)
        out = tmp_path / "depletion.csv"

        result_path = write_stream_depletion_csv(report, out)

        assert result_path == out
        with out.open(encoding="utf-8") as f:
            rows = list(csv.reader(f))

        # Golden header — column order is part of the public CSV contract
        assert rows[0] == [
            "reach_id",
            "reach_name",
            "time",
            "baseline_flow",
            "scenario_flow",
            "depletion",
            "cumulative_depletion",
        ]

        # 2 reaches * 3 timesteps = 6 data rows
        data = rows[1:]
        assert len(data) == 6
        # First row: reach 1, first timestep
        assert data[0][0] == "1"
        assert data[0][1] == "Reach 1"
        assert data[0][2] == "2024-01-01"

    def test_round_trip_values(self, tmp_path: Path):
        report = _make_report(n_reaches=1, n_timesteps=2)
        out = tmp_path / "out.csv"
        write_stream_depletion_csv(report, out)

        with out.open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        r0 = report.results[0]
        # Values are formatted with %.6g (6 significant figures); allow that
        # rounding window in the comparison.
        rel = 1e-5
        assert float(rows[0]["baseline_flow"]) == pytest.approx(r0.baseline_flow[0], rel=rel)
        assert float(rows[0]["depletion"]) == pytest.approx(r0.depletion[0], rel=rel)
        assert float(rows[1]["cumulative_depletion"]) == pytest.approx(
            r0.cumulative_depletion[1], rel=rel
        )

    def test_accepts_string_path(self, tmp_path: Path):
        report = _make_report(n_reaches=1, n_timesteps=1)
        out = str(tmp_path / "out.csv")
        result = write_stream_depletion_csv(report, out)
        assert result == Path(out)
        assert Path(out).exists()


class TestJsonWriter:
    def test_round_trip_via_to_dict(self, tmp_path: Path):
        report = _make_report(n_reaches=2, n_timesteps=3)
        out = tmp_path / "depletion.json"

        write_stream_depletion_json(report, out)

        loaded = json.loads(out.read_text())
        assert loaded == report.to_dict()
        assert loaded["n_reaches"] == 2
        assert len(loaded["reaches"]) == 2

    def test_indent_default_is_2(self, tmp_path: Path):
        report = _make_report(n_reaches=1, n_timesteps=1)
        out = tmp_path / "out.json"
        write_stream_depletion_json(report, out)
        text = out.read_text()
        # Pretty-printed: should have multiple lines and 2-space indent
        assert "\n" in text
        assert "  " in text

    def test_indent_none_compact(self, tmp_path: Path):
        report = _make_report(n_reaches=1, n_timesteps=1)
        out = tmp_path / "out.json"
        write_stream_depletion_json(report, out, indent=None)
        text = out.read_text()
        # Compact: no newlines
        assert "\n" not in text


class TestExcelWriter:
    def test_summary_sheet(self, tmp_path: Path):
        report = _make_report(n_reaches=2, n_timesteps=3)
        out = tmp_path / "depletion.xlsx"

        write_stream_depletion_excel(report, out)

        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames

        summary = wb["Summary"]
        rows = list(summary.values)
        assert rows[0] == (
            "reach_id",
            "reach_name",
            "max_depletion",
            "max_depletion_timestep",
            "total_depletion",
        )
        # n_reaches data rows + 1 blank + 1 totals = 4
        assert rows[1][0] == 1
        assert rows[2][0] == 2
        # Totals row is last
        assert rows[-1][0] == "TOTAL"

    def test_one_sheet_per_reach(self, tmp_path: Path):
        report = _make_report(n_reaches=3, n_timesteps=2)
        out = tmp_path / "depletion.xlsx"

        write_stream_depletion_excel(report, out)

        wb = load_workbook(out, read_only=True)
        # Summary + 3 reach sheets
        assert len(wb.sheetnames) == 4

    def test_invalid_sheet_chars_sanitized(self, tmp_path: Path):
        # Put illegal Excel chars in the reach name
        report = _make_report(n_reaches=1, n_timesteps=1)
        report.results[0].reach_name = "Reach/1:foo*?"
        out = tmp_path / "out.xlsx"

        write_stream_depletion_excel(report, out)

        wb = load_workbook(out, read_only=True)
        # Original chars stripped; name must not contain any of [/\:*?[]
        for name in wb.sheetnames:
            for ch in r"[]:*?/\\":
                assert ch not in name

    def test_long_reach_name_truncated(self, tmp_path: Path):
        report = _make_report(n_reaches=1, n_timesteps=1)
        report.results[0].reach_name = "A" * 100  # > 31 chars
        out = tmp_path / "out.xlsx"
        write_stream_depletion_excel(report, out)
        wb = load_workbook(out, read_only=True)
        for name in wb.sheetnames:
            assert len(name) <= 31


class TestReportWriteDispatcher:
    def test_dispatch_by_csv_extension(self, tmp_path: Path):
        report = _make_report()
        out = tmp_path / "out.csv"
        report.write(out)
        # CSV header should be present
        first_line = out.read_text().splitlines()[0]
        assert first_line.startswith("reach_id,")

    def test_dispatch_by_json_extension(self, tmp_path: Path):
        report = _make_report()
        out = tmp_path / "out.json"
        report.write(out)
        loaded = json.loads(out.read_text())
        assert "reaches" in loaded

    def test_dispatch_by_xlsx_extension(self, tmp_path: Path):
        report = _make_report()
        out = tmp_path / "out.xlsx"
        report.write(out)
        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames

    def test_explicit_format_overrides_extension(self, tmp_path: Path):
        report = _make_report()
        # File extension says .json but caller explicitly asks for csv
        out = tmp_path / "out.json"
        report.write(out, format="csv")
        first_line = out.read_text().splitlines()[0]
        assert first_line.startswith("reach_id,")

    def test_unknown_extension_raises(self, tmp_path: Path):
        report = _make_report()
        with pytest.raises(ValueError, match=r"expected one of: csv, json, xlsx"):
            report.write(tmp_path / "out.unknown")

    def test_unknown_explicit_format_raises(self, tmp_path: Path):
        report = _make_report()
        with pytest.raises(ValueError, match=r"Unknown format"):
            report.write(tmp_path / "out.csv", format="parquet")

    def test_excel_alias(self, tmp_path: Path):
        # ``format="excel"`` should work as an alias for "xlsx"
        report = _make_report()
        out = tmp_path / "out.xlsx"
        report.write(out, format="excel")
        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames
