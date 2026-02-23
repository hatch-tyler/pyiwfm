"""Tests for pyiwfm.io.budget_excel — budget Excel export."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from pyiwfm.io.budget import ASCIIOutputInfo, BudgetHeader, LocationData, TimeStepInfo
from pyiwfm.io.budget_excel import budget_to_excel


def _make_mock_reader(
    n_locations: int = 2,
    n_timesteps: int = 5,
    n_columns: int = 3,
    descriptor: str = "GROUNDWATER BUDGET",
) -> MagicMock:
    """Create a mock BudgetReader for testing."""
    reader = MagicMock()
    reader.n_locations = n_locations
    reader.n_timesteps = n_timesteps

    loc_names = [f"Subregion {i + 1}" for i in range(n_locations)]
    reader.locations = loc_names

    # Header
    header = BudgetHeader()
    header.descriptor = descriptor
    header.n_locations = n_locations
    header.location_names = loc_names
    header.areas = np.array([1000.0 * (i + 1) for i in range(n_locations)])
    header.n_areas = n_locations

    ts = TimeStepInfo()
    ts.n_timesteps = n_timesteps
    ts.start_datetime = datetime(2000, 1, 1)
    ts.delta_t_minutes = 43200  # ~30 days
    ts.unit = "1MON"
    header.timestep = ts

    ascii_out = ASCIIOutputInfo()
    ascii_out.titles = [
        f"{descriptor}",
        "FOR @LOCNAME@",
        "VOLUME = @UNITVL@, AREA = @UNITAR@",
    ]
    ascii_out.n_titles = len(ascii_out.titles)
    header.ascii_output = ascii_out

    col_headers = [f"Column {j + 1}" for j in range(n_columns)]
    loc_data = LocationData()
    loc_data.n_columns = n_columns
    loc_data.column_headers = col_headers
    loc_data.column_types = [1, 4, 5][:n_columns]  # VR, AR, LT
    loc_data.column_widths = [15] * n_columns
    header.location_data = [loc_data]

    reader.header = header

    def _get_column_headers(loc: str | int = 0) -> list[str]:
        return col_headers

    reader.get_column_headers = _get_column_headers

    def _get_location_index(loc: str | int) -> int:
        if isinstance(loc, int):
            return loc
        return loc_names.index(loc)

    reader.get_location_index = _get_location_index

    def _get_dataframe(
        loc: str | int = 0,
        columns: object = None,
        *,
        length_factor: float = 1.0,
        area_factor: float = 1.0,
        volume_factor: float = 1.0,
    ) -> pd.DataFrame:
        idx = _get_location_index(loc)
        dates = pd.date_range("2000-01-01", periods=n_timesteps, freq="MS")
        data = np.ones((n_timesteps, n_columns)) * (idx + 1)
        # Apply conversion factors per column type (mimics real get_dataframe)
        col_types = [1, 4, 5][:n_columns]  # VR, AR, LT
        factor_map = {1: volume_factor, 4: area_factor, 5: length_factor}
        for j, ct in enumerate(col_types):
            data[:, j] *= factor_map.get(ct, 1.0)
        return pd.DataFrame(data, index=dates, columns=col_headers)

    reader.get_dataframe = _get_dataframe

    return reader


class TestBudgetToExcel:
    """Tests for budget_to_excel()."""

    def test_creates_xlsx(self, tmp_path: Path) -> None:
        reader = _make_mock_reader()
        out = tmp_path / "test.xlsx"
        result = budget_to_excel(reader, out)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_one_sheet_per_location(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_locations=3)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 3

    def test_sheet_names(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_locations=2)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out)
        wb = load_workbook(out)
        assert "Subregion 1" in wb.sheetnames
        assert "Subregion 2" in wb.sheetnames

    def test_title_lines_bold(self, tmp_path: Path) -> None:
        reader = _make_mock_reader()
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # First few rows should be titles (bold)
        assert ws.cell(row=1, column=1).font.bold is True

    def test_title_substitution(self, tmp_path: Path) -> None:
        reader = _make_mock_reader()
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out, volume_unit="TAF", area_unit="ACRES")
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # Title line 3 should have substituted units
        title3 = ws.cell(row=3, column=1).value
        assert "TAF" in title3
        assert "ACRES" in title3

    def test_header_row_bold(self, tmp_path: Path) -> None:
        reader = _make_mock_reader()
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # Row 5 should be headers (3 titles + 1 blank + 1 header)
        header_row = 5
        assert ws.cell(row=header_row, column=1).font.bold is True
        assert ws.cell(row=header_row, column=1).value == "Time"

    def test_data_rows_present(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_timesteps=5)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # Data starts at row 6 (3 titles + 1 blank + 1 header)
        data_start = 6
        assert ws.cell(row=data_start, column=1).value is not None

    def test_location_filter(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_locations=5)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out, location_ids=[1, 3])
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2

    def test_unit_conversion_applied(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_columns=3)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out, volume_factor=2.0, area_factor=3.0, length_factor=4.0)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        data_start = 6
        # Column 2 (VR type=1) should be multiplied by volume_factor=2.0
        # Original value is 1.0 (mock returns ones * (loc_idx+1) = 1)
        val_col1 = ws.cell(row=data_start, column=2).value
        assert val_col1 == pytest.approx(2.0, abs=0.01)
        # Column 3 (AR type=4) → area_factor=3.0
        val_col2 = ws.cell(row=data_start, column=3).value
        assert val_col2 == pytest.approx(3.0, abs=0.01)
        # Column 4 (LT type=5) → length_factor=4.0
        val_col3 = ws.cell(row=data_start, column=4).value
        assert val_col3 == pytest.approx(4.0, abs=0.01)

    def test_empty_locations_placeholder(self, tmp_path: Path) -> None:
        reader = _make_mock_reader(n_locations=1)
        out = tmp_path / "test.xlsx"
        budget_to_excel(reader, out, location_ids=[])
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "(no data)"

    def test_creates_parent_dirs(self, tmp_path: Path) -> None:
        reader = _make_mock_reader()
        out = tmp_path / "sub" / "dir" / "test.xlsx"
        result = budget_to_excel(reader, out)
        assert result.exists()
