"""Tests for the Phase 2 drawdown report dataclasses, builders, and
tabular writers.

Uses a mock :class:`LazyNodalLoader` so the tests don't need real
HDF5 files. The mock returns deterministic synthetic head fields so we
can verify the drawdown arithmetic.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock

import numpy as np
import pytest
from openpyxl import load_workbook

from pyiwfm.io.drawdown import (
    DrawdownAtLocation,
    DrawdownComputer,
    DrawdownSnapshot,
    DrawdownTimeSeriesReport,
    write_drawdown_timeseries_csv,
    write_drawdown_timeseries_excel,
    write_drawdown_timeseries_json,
)


def _make_loader(
    n_nodes: int = 5,
    n_layers: int = 2,
    n_timesteps: int = 4,
) -> MagicMock:
    """Build a mock LazyNodalLoader that returns synthetic heads.

    Synthetic head field: head[t, node, layer] = 100 - t * (node + layer)
    so drawdown(t, layer) at node n vs reference t=0 is t * (n + layer).
    Predictable and easy to assert against.
    """
    loader = MagicMock()
    loader.n_nodes = n_nodes
    loader.n_layers = n_layers
    loader.n_frames = n_timesteps
    base = datetime(2024, 1, 1)
    loader.times = [base + timedelta(days=t) for t in range(n_timesteps)]

    def _get_frame(t: int) -> np.ndarray:
        # shape: (n_nodes, n_layers)
        nodes = np.arange(1, n_nodes + 1)[:, None]  # (n_nodes, 1)
        layers = np.arange(1, n_layers + 1)[None, :]  # (1, n_layers)
        return 100.0 - float(t) * (nodes + layers)

    loader.get_frame.side_effect = _get_frame
    return loader


class TestBuildTimeseriesReport:
    def test_arithmetic_at_one_location(self):
        loader = _make_loader(n_nodes=5, n_layers=2, n_timesteps=4)
        computer = DrawdownComputer(loader)

        report = computer.build_timeseries_report(locations=[(3, 1)])

        assert isinstance(report, DrawdownTimeSeriesReport)
        assert report.n_locations == 1
        assert report.n_timesteps == 4
        assert report.reference_timestep == 0
        loc = report.locations[0]
        assert loc.node_id == 3
        assert loc.layer == 1
        # Synthetic heads: head[t, n=3, layer=1] = 100 - t * (3+1) = 100 - 4t
        # drawdown vs t=0 = head(0) - head(t) = 100 - (100 - 4t) = 4t
        np.testing.assert_array_almost_equal(loc.drawdown, [0.0, 4.0, 8.0, 12.0])
        assert loc.max_drawdown == pytest.approx(12.0)
        assert loc.max_drawdown_timestep == 3
        assert loc.final_drawdown == pytest.approx(12.0)

    def test_multiple_locations_share_times(self):
        loader = _make_loader(n_nodes=5, n_layers=2, n_timesteps=3)
        computer = DrawdownComputer(loader)

        report = computer.build_timeseries_report(
            locations=[(1, 1), (5, 2)],
        )

        assert report.n_locations == 2
        # All locations share the same timestamp list
        assert report.times == report.locations[0].times == report.locations[1].times
        # Values differ by location
        # node=1, layer=1: head(t) = 100 - t*2; drawdown = 2t -> [0, 2, 4]
        np.testing.assert_array_almost_equal(report.locations[0].drawdown, [0, 2, 4])
        # node=5, layer=2: head(t) = 100 - t*7; drawdown = 7t -> [0, 7, 14]
        np.testing.assert_array_almost_equal(report.locations[1].drawdown, [0, 7, 14])

    def test_custom_reference_timestep(self):
        loader = _make_loader(n_nodes=3, n_layers=1, n_timesteps=4)
        computer = DrawdownComputer(loader)

        # Reference at t=2 instead of t=0
        report = computer.build_timeseries_report(
            locations=[(2, 1)],
            reference_timestep=2,
        )
        assert report.reference_timestep == 2
        # head(t, n=2, l=1) = 100 - 3t. ref=t=2 -> head_ref = 94.
        # drawdown = 94 - (100 - 3t) = 3t - 6 -> [-6, -3, 0, 3]
        np.testing.assert_array_almost_equal(report.locations[0].drawdown, [-6, -3, 0, 3])
        # max_drawdown is the largest absolute value -> -6 at index 0
        assert abs(report.locations[0].max_drawdown) == pytest.approx(6.0)

    def test_empty_locations_returns_empty_report(self):
        loader = _make_loader()
        computer = DrawdownComputer(loader)
        report = computer.build_timeseries_report(locations=[])
        assert report.n_locations == 0
        assert report.locations == []

    def test_invalid_node_id_raises(self):
        loader = _make_loader(n_nodes=3, n_layers=1, n_timesteps=2)
        computer = DrawdownComputer(loader)
        with pytest.raises(IndexError, match=r"node_id 99 out of range \[1, 3\]"):
            computer.build_timeseries_report(locations=[(99, 1)])

    def test_invalid_layer_raises(self):
        loader = _make_loader(n_nodes=3, n_layers=2, n_timesteps=2)
        computer = DrawdownComputer(loader)
        with pytest.raises(IndexError, match=r"layer 5 out of range \[1, 2\]"):
            computer.build_timeseries_report(locations=[(1, 5)])

    def test_invalid_reference_timestep_raises(self):
        loader = _make_loader(n_timesteps=3)
        computer = DrawdownComputer(loader)
        with pytest.raises(IndexError, match="Timestep 99 out of range"):
            computer.build_timeseries_report(locations=[(1, 1)], reference_timestep=99)


class TestBuildSnapshot:
    def test_snapshot_has_one_value_per_node(self):
        loader = _make_loader(n_nodes=5, n_layers=2, n_timesteps=3)
        computer = DrawdownComputer(loader)

        snap = computer.build_snapshot(timestep=2, layer=1, reference_timestep=0)

        assert isinstance(snap, DrawdownSnapshot)
        assert snap.n_nodes == 5
        assert snap.kind == "single"
        assert snap.timestep == 2
        assert snap.layer == 1
        # node IDs are 1-based
        assert snap.node_ids.tolist() == [1, 2, 3, 4, 5]
        # head[t=2, n, l=1] = 100 - 2*(n+1); ref t=0 -> 100. drawdown = 2*(n+1)
        np.testing.assert_array_almost_equal(snap.drawdown, [4, 6, 8, 10, 12])

    def test_max_snapshot_is_per_node_max_across_time(self):
        loader = _make_loader(n_nodes=4, n_layers=1, n_timesteps=5)
        computer = DrawdownComputer(loader)

        snap = computer.build_max_snapshot(layer=1)

        assert snap.kind == "max"
        assert snap.timestep == -1
        # head(t, n, l=1) = 100 - t*(n+1); drawdown(t, n) = t*(n+1)
        # max across t=0..4 for each n: 4*(n+1) -> [8, 12, 16, 20]
        np.testing.assert_array_almost_equal(snap.drawdown, [8, 12, 16, 20])

    def test_time_label_formatted(self):
        loader = _make_loader(n_timesteps=3)
        computer = DrawdownComputer(loader)
        snap = computer.build_snapshot(timestep=1, layer=1)
        # The synthetic loader uses Jan 2-Jan 3 timestamps
        assert snap.time_label.startswith("2024-01-02")


class TestDataclassToDict:
    def test_drawdown_at_location_to_dict(self):
        loc = DrawdownAtLocation(
            node_id=1,
            layer=2,
            times=["2024-01-01", "2024-01-02"],
            drawdown=np.array([1.0, 2.0]),
            max_drawdown=2.0,
            max_drawdown_timestep=1,
            final_drawdown=2.0,
        )
        d = loc.to_dict()
        assert d["node_id"] == 1
        assert d["layer"] == 2
        assert d["drawdown"] == [1.0, 2.0]
        assert d["max_drawdown"] == 2.0

    def test_to_dict_handles_nan(self):
        loc = DrawdownAtLocation(
            node_id=1,
            layer=1,
            times=["t0"],
            drawdown=np.array([np.nan]),
            max_drawdown=float("nan"),
            max_drawdown_timestep=0,
            final_drawdown=float("nan"),
        )
        d = loc.to_dict()
        assert d["drawdown"] == [None]
        assert d["max_drawdown"] is None
        assert d["final_drawdown"] is None

    def test_snapshot_to_dict(self):
        snap = DrawdownSnapshot(
            timestep=2,
            layer=1,
            reference_timestep=0,
            kind="single",
            time_label="2024-01-03",
            node_ids=np.array([1, 2, 3], dtype=np.int32),
            drawdown=np.array([1.0, np.nan, 3.0]),
            n_nodes=3,
        )
        d = snap.to_dict()
        assert d["timestep"] == 2
        assert d["kind"] == "single"
        assert d["node_ids"] == [1, 2, 3]
        # NaN -> null
        assert d["drawdown"] == [1.0, None, 3.0]


def _build_small_report(n_locations: int = 2) -> DrawdownTimeSeriesReport:
    loader = _make_loader(n_nodes=5, n_layers=2, n_timesteps=3)
    computer = DrawdownComputer(loader)
    return computer.build_timeseries_report(
        locations=[(i, 1) for i in range(1, n_locations + 1)],
    )


class TestCsvWriter:
    def test_writes_long_format(self, tmp_path: Path):
        report = _build_small_report(n_locations=2)
        out = tmp_path / "drawdown.csv"

        result = write_drawdown_timeseries_csv(report, out)

        assert result == out
        with out.open(encoding="utf-8") as f:
            rows = list(csv.reader(f))

        # Golden header
        assert rows[0] == ["node_id", "layer", "time", "drawdown"]
        # 2 locations × 3 timesteps = 6 data rows
        assert len(rows) - 1 == 6
        # First data row: location 1, timestep 0
        assert rows[1][0] == "1"
        assert rows[1][1] == "1"

    def test_nan_emitted_as_empty(self, tmp_path: Path):
        report = _build_small_report(n_locations=1)
        # Inject a NaN
        report.locations[0].drawdown[1] = np.nan
        out = tmp_path / "out.csv"
        write_drawdown_timeseries_csv(report, out)
        with out.open(encoding="utf-8") as f:
            rows = list(csv.reader(f))
        # Row for the NaN timestep has empty drawdown
        assert rows[2][3] == ""


class TestJsonWriter:
    def test_round_trip_via_to_dict(self, tmp_path: Path):
        report = _build_small_report(n_locations=2)
        out = tmp_path / "out.json"
        write_drawdown_timeseries_json(report, out)
        loaded = json.loads(out.read_text())
        assert loaded == report.to_dict()
        assert loaded["n_locations"] == 2

    def test_compact_when_indent_none(self, tmp_path: Path):
        report = _build_small_report(n_locations=1)
        out = tmp_path / "out.json"
        write_drawdown_timeseries_json(report, out, indent=None)
        text = out.read_text()
        assert "\n" not in text


class TestExcelWriter:
    def test_summary_plus_per_location_sheets(self, tmp_path: Path):
        report = _build_small_report(n_locations=3)
        out = tmp_path / "drawdown.xlsx"

        write_drawdown_timeseries_excel(report, out)

        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames
        # Summary + 3 location sheets
        assert len(wb.sheetnames) == 4
        # Sheet names follow the Node{id}_L{layer} convention
        loc_sheets = [s for s in wb.sheetnames if s != "Summary"]
        assert any("Node1_L1" == s for s in loc_sheets)

    def test_summary_columns(self, tmp_path: Path):
        report = _build_small_report(n_locations=1)
        out = tmp_path / "out.xlsx"
        write_drawdown_timeseries_excel(report, out)
        wb = load_workbook(out, read_only=True)
        summary = wb["Summary"]
        rows = list(summary.values)
        assert rows[0] == (
            "node_id",
            "layer",
            "max_drawdown",
            "max_drawdown_timestep",
            "final_drawdown",
        )
        assert rows[1][0] == 1

    def test_long_label_truncated(self, tmp_path: Path):
        report = _build_small_report(n_locations=1)
        # Force a node_id that produces a >31-char sheet name when combined
        # with the layer suffix. With node_id huge, "Node{id}_L1" can
        # exceed 31; verify we truncate.
        report.locations[0].node_id = 99999999999999999999999  # type: ignore[assignment]
        out = tmp_path / "out.xlsx"
        write_drawdown_timeseries_excel(report, out)
        wb = load_workbook(out, read_only=True)
        for name in wb.sheetnames:
            assert len(name) <= 31


class TestReportWriteDispatcher:
    def test_dispatch_by_csv_extension(self, tmp_path: Path):
        report = _build_small_report()
        out = tmp_path / "out.csv"
        report.write(out)  # type: ignore[attr-defined]
        first_line = out.read_text().splitlines()[0]
        assert first_line.startswith("node_id,")

    def test_dispatch_by_json_extension(self, tmp_path: Path):
        report = _build_small_report()
        out = tmp_path / "out.json"
        report.write(out)  # type: ignore[attr-defined]
        loaded = json.loads(out.read_text())
        assert "locations" in loaded

    def test_dispatch_by_xlsx_extension(self, tmp_path: Path):
        report = _build_small_report()
        out = tmp_path / "out.xlsx"
        report.write(out)  # type: ignore[attr-defined]
        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames

    def test_excel_alias(self, tmp_path: Path):
        report = _build_small_report()
        out = tmp_path / "out.xlsx"
        report.write(out, format="excel")  # type: ignore[attr-defined]
        wb = load_workbook(out, read_only=True)
        assert "Summary" in wb.sheetnames

    def test_unknown_extension_raises(self, tmp_path: Path):
        report = _build_small_report()
        with pytest.raises(ValueError, match=r"expected one of: csv, json, xlsx"):
            report.write(tmp_path / "out.txt")  # type: ignore[attr-defined]

    def test_explicit_format_overrides_extension(self, tmp_path: Path):
        report = _build_small_report()
        out = tmp_path / "out.json"  # extension says JSON
        report.write(out, format="csv")  # type: ignore[attr-defined]
        first_line = out.read_text().splitlines()[0]
        assert first_line.startswith("node_id,")
