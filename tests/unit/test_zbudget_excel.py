"""Tests for pyiwfm.io.zbudget_excel — zone budget Excel export."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from pyiwfm.io.zbudget import ZBudgetHeader, ZoneInfo
from pyiwfm.io.zbudget_excel import zbudget_to_excel


def _make_mock_zreader(
    n_zones: int = 2,
    n_timesteps: int = 5,
    n_data: int = 3,
    descriptor: str = "GROUNDWATER ZONE BUDGET",
) -> MagicMock:
    """Create a mock ZBudgetReader for testing."""
    reader = MagicMock()

    zone_names = [f"Zone {i + 1}" for i in range(n_zones)]
    reader.zones = zone_names
    reader.n_zones = n_zones
    reader.n_timesteps = n_timesteps
    reader.descriptor = descriptor
    reader.data_names = [f"Data {j + 1}" for j in range(n_data)]

    header = ZBudgetHeader()
    header.descriptor = descriptor
    header.n_timesteps = n_timesteps
    header.start_datetime = datetime(2000, 1, 1)
    header.delta_t_minutes = 43200
    header.time_unit = "1MON"
    header.data_names = reader.data_names
    header.n_data = n_data
    reader.header = header

    def _get_zone_info(zone: str | int) -> ZoneInfo:
        if isinstance(zone, int):
            idx = zone - 1
        else:
            idx = zone_names.index(zone)
        return ZoneInfo(id=idx + 1, name=zone_names[idx], area=5000.0 * (idx + 1))

    reader.get_zone_info = _get_zone_info

    def _get_dataframe(
        zone: str | int,
        layer: int = 1,
        data_columns: list[str] | None = None,
        *,
        volume_factor: float = 1.0,
    ) -> pd.DataFrame:
        if isinstance(zone, int):
            z_idx = zone - 1
        else:
            z_idx = zone_names.index(zone)
        dates = pd.date_range("2000-01-01", periods=n_timesteps, freq="MS")
        data = np.ones((n_timesteps, n_data)) * (z_idx + 1) * volume_factor
        df = pd.DataFrame(data, index=dates, columns=reader.data_names)
        if data_columns is not None:
            df = df[[c for c in data_columns if c in df.columns]]
        return df

    reader.get_dataframe = _get_dataframe

    return reader


class TestZBudgetToExcel:
    """Tests for zbudget_to_excel()."""

    def test_creates_xlsx(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader()
        out = tmp_path / "test.xlsx"
        result = zbudget_to_excel(reader, out)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_one_sheet_per_zone(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader(n_zones=3)
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 3

    def test_sheet_names(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader(n_zones=2)
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out)
        wb = load_workbook(out)
        assert "Zone 1" in wb.sheetnames
        assert "Zone 2" in wb.sheetnames

    def test_title_bold(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader()
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        assert ws.cell(row=1, column=1).font.bold is True

    def test_volume_conversion(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader(n_zones=1)
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out, volume_factor=10.0)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # Find data start (after title + area + unit + blank + header rows)
        data_start = 6
        val = ws.cell(row=data_start, column=2).value
        assert val == pytest.approx(10.0, abs=0.01)

    def test_zone_filter(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader(n_zones=5)
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out, zone_ids=[2, 4])
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2

    def test_empty_zones_placeholder(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader()
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out, zone_ids=[])
        wb = load_workbook(out)
        assert wb.sheetnames[0] == "(no data)"

    def test_data_rows_present(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader(n_timesteps=3)
        out = tmp_path / "test.xlsx"
        zbudget_to_excel(reader, out)
        wb = load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        # Should have 3 data rows below the header
        data_start = 6
        assert ws.cell(row=data_start, column=1).value is not None
        assert ws.cell(row=data_start + 2, column=1).value is not None

    def test_creates_parent_dirs(self, tmp_path: Path) -> None:
        reader = _make_mock_zreader()
        out = tmp_path / "nested" / "dir" / "test.xlsx"
        result = zbudget_to_excel(reader, out)
        assert result.exists()
